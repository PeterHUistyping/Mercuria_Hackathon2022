
# Python program to read an excel file 
# import openpyxl module 
import openpyxl 
from statistics import mean
# Give the location of the file 
path = "./Excel/hercules.xlsx"
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
# print("Total Rows:", row)
# print("Total Columns:", column)
# printing the value of first column
# Loop will print all values 
# of first column  
# print("\nValue of first column")

timestamp = []
velocity = []
distance=[]
fuel=[]
ghg=[]

start_row=207
for i in range(207, 231+1): 
    ts_obj = sheet_obj.cell(row = i, column = 1) 
    # if type(ts_obj) is None:
    if i==215 or i==228 :
        continue
    print()
    timestamp.append(ts_obj.value)
    velocity_obj = sheet_obj.cell(row = i, column = 4) 
    vel=(float)(velocity_obj.value)
    velocity.append(vel)
    distance.append(24*vel)
    fuel_eachpart=[]
    # sum=0
    for j in range(15,20):
        fuel_obj = sheet_obj.cell(row = i, column = j) 
        fuel_eachpart.append(1000*fuel_obj.value) #putting into Tons instead of kT
        # print(fuel_obj.value)
        # sum+= fuel_obj.value
    # sum/=6.0
    fuel.append(sum(fuel_eachpart))
    # distance = sheet_obj.cell(row = i, column = 4)  24*velocity
    if(i>=215):
        start_row=208
    if(i>=228):
        start_row=209
    ghg.append(3.206*fuel[i-start_row]) # normalised to kT
    print(timestamp[i-start_row],velocity[i-start_row],distance[i-start_row],fuel[i-start_row],ghg[i-start_row])    

# printing the value of first column
# Loop will print all values 
# of first row
# print("\nValue of first row")
# for i in range(1, column + 1): 
#     cell_obj = sheet_obj.cell(row = 2, column = i) 
#     print(cell_obj.value, end = " ")