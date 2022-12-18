
# Python program to read an excel file 
# import openpyxl module 
import openpyxl 
from statistics import mean
# Give the location of the file 
path = "./Port/mar.xlsx"
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
# print("Total Rows:", row)
# print("Total Columns:", column)
# printing the value of first column
# Loop will print all values 
# of first column  
# print("\nValue of first column")
country_name=[]
lat = []
lon = []
ports = {}
end_row=500 # max row+1
for i in range(2, end_row): 
    lon_obj = sheet_obj.cell(row = i, column = 1) 
    lon_value=(lon_obj.value)
    lat_obj = sheet_obj.cell(row = i, column = 2) 
    lat_value=(lat_obj.value)
    if lat_value =='' or lon_value =='' or lat_value is None or lon_value is None:
        continue
    con_obj = sheet_obj.cell(row = i, column = 4) 
    country_value=con_obj.value
    # if type(ts_obj) is None:
    # country_name.append(country_value)
    # lat.append(float(lat_value))
    # lon.append(float(lon_value))
    print(con_obj.value,lat_value,lon_value)
    ports[country_value] = {"lat" :float(lat_value) , "lon" : float(lon_value)}
print(ports)

