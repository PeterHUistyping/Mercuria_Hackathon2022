
# Python program to read an excel file 
# import openpyxl module 
import openpyxl 
from statistics import mean
# Give the location of the file 
path = "./Port/mar.xlsx"
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
# print("Total Rows:", row)
# print("Total Columns:", column)
# printing the value of first column
# Loop will print all values 
# of first column  
# print("\nValue of first column")
country_name=[]
lat = []
lon = []
start_row=2
for i in range(2, row+1): 
    lat_obj = sheet_obj.cell(row = i, column = 1) 
    lat_value=(lat_obj.value)
    lon_obj = sheet_obj.cell(row = i, column = 2) 
    lon_value=(lon_obj.value)
    if lat_value =='' or lon_value =='' or lat_value is None or lon_value is None:
        continue
    con_obj = sheet_obj.cell(row = i, column = 4) 
    # if type(ts_obj) is None:
    country_name.append(con_obj.value)
    lat.append(lat_value)
    lon.append(lon_value)
    print(con_obj.value,lat_value,lon_value)